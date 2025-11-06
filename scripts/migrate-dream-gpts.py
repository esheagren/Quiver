#!/usr/bin/env python3
"""
DREAM GPT Database Migration Script
Migrates top GPTs from Excel database to Quiver Supabase
"""

import openpyxl
import os
from datetime import datetime
from supabase import create_client, Client
from dotenv import load_dotenv
import uuid

# Load environment variables
load_dotenv()

SUPABASE_URL = os.getenv('VITE_SUPABASE_URL')
SUPABASE_KEY = os.getenv('VITE_SUPABASE_ANON_KEY')
EXCEL_FILE = '/Users/erik/Downloads/DREAM Centralized GPT Database (1).xlsx'

# Migration user token (consistent UUID for all migrated GPTs)
MIGRATION_USER_TOKEN = 'a1b2c3d4-e5f6-4a7b-8c9d-0e1f2a3b4c5d'


def init_supabase() -> Client:
    """Initialize Supabase client"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise ValueError("Missing Supabase credentials in .env file")
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def clear_existing_data(supabase: Client):
    """Clear all existing prompt data from Supabase"""
    print("🗑️  Clearing existing data...")

    # Delete in order due to foreign key constraints
    # 1. Delete upvotes first (references prompts)
    result = supabase.table('prompt_upvotes').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    print(f"   Deleted upvotes")

    # 2. Delete saved prompts (references prompts)
    result = supabase.table('saved_prompts').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    print(f"   Deleted saved prompts")

    # 3. Delete prompts
    result = supabase.table('prompts').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    print(f"   Deleted prompts")

    print("✅ All existing data cleared!\n")


def read_excel_and_get_top_gpts(limit: int = 5):
    """Read Excel file and extract top GPTs by Average Score"""
    print(f"📊 Reading Excel file: {EXCEL_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Centralized Database']

    # Column indices (1-based)
    COL_GPT_ID = 1
    COL_GPT_NAME = 2
    COL_DESCRIPTION = 3
    COL_CREATOR_EMAIL = 4
    COL_TAGS = 5
    COL_FIRST_ACTIVE = 6
    COL_LAST_ACTIVE = 7
    COL_TOTAL_MESSAGES = 10
    COL_UNIQUE_USERS = 12
    COL_AVERAGE_SCORE = 16

    # Read all GPTs
    gpts = []
    for row_num in range(2, ws.max_row + 1):
        gpt_id = ws.cell(row_num, COL_GPT_ID).value
        if not gpt_id:
            break

        gpt_data = {
            'gpt_id': str(gpt_id),
            'name': ws.cell(row_num, COL_GPT_NAME).value or 'Untitled GPT',
            'description': ws.cell(row_num, COL_DESCRIPTION).value or 'No description available',
            'creator_email': ws.cell(row_num, COL_CREATOR_EMAIL).value,
            'tags': ws.cell(row_num, COL_TAGS).value,
            'first_active': ws.cell(row_num, COL_FIRST_ACTIVE).value,
            'last_active': ws.cell(row_num, COL_LAST_ACTIVE).value,
            'total_messages': ws.cell(row_num, COL_TOTAL_MESSAGES).value or 0,
            'unique_users': ws.cell(row_num, COL_UNIQUE_USERS).value or 0,
            'average_score': ws.cell(row_num, COL_AVERAGE_SCORE).value or 0,
        }
        gpts.append(gpt_data)

    # Sort by Average Score (descending)
    gpts.sort(key=lambda x: float(x['average_score']) if x['average_score'] else 0, reverse=True)

    # Get top N
    top_gpts = gpts[:limit]

    print(f"✅ Found {len(gpts)} total GPTs")
    print(f"📌 Selected top {limit} by Average Score:\n")

    for i, gpt in enumerate(top_gpts, 1):
        print(f"   {i}. {gpt['name']}")
        print(f"      Score: {gpt['average_score']}, Users: {gpt['unique_users']}, Messages: {gpt['total_messages']}")

    print()
    return top_gpts


def parse_tags(tags_str: str) -> list:
    """Parse semicolon-separated tags string into list"""
    if not tags_str:
        return ['Other']

    tags = [t.strip() for t in str(tags_str).split(';') if t.strip()]
    return tags if tags else ['Other']


def format_date(date_value) -> str:
    """Convert Excel date to ISO format string"""
    if not date_value:
        return datetime.now().isoformat()

    if isinstance(date_value, datetime):
        return date_value.isoformat()

    # Try to parse string date
    try:
        dt = datetime.strptime(str(date_value), '%Y-%m-%d')
        return dt.isoformat()
    except:
        return datetime.now().isoformat()


def insert_gpts_to_supabase(supabase: Client, gpts: list):
    """Insert GPTs into Supabase prompts table"""
    print("💾 Inserting GPTs into Supabase...\n")

    inserted_count = 0

    for gpt in gpts:
        # Construct URL from GPT ID
        url = f"https://chatgpt.com/g/{gpt['gpt_id']}"

        # Parse tags
        tags = parse_tags(gpt['tags'])

        # Prepare prompt data
        prompt_data = {
            'url': url,
            'prompt_text': gpt['description'],  # Using description as prompt_text placeholder
            'generated_name': gpt['name'],
            'description': gpt['description'],
            'tags': tags,
            'user_token': MIGRATION_USER_TOKEN,
            'upvotes': int(gpt['unique_users']),  # Unique users as upvotes
            'uses': int(gpt['total_messages']),   # Total messages as uses
            'created_at': format_date(gpt['first_active']),
        }

        try:
            result = supabase.table('prompts').insert(prompt_data).execute()
            inserted_count += 1

            print(f"✅ Inserted: {gpt['name']}")
            print(f"   URL: {url}")
            print(f"   Tags: {', '.join(tags)}")
            print(f"   Upvotes: {prompt_data['upvotes']}, Uses: {prompt_data['uses']}")
            print()

        except Exception as e:
            print(f"❌ Error inserting {gpt['name']}: {e}")
            print()

    print(f"🎉 Successfully inserted {inserted_count}/{len(gpts)} GPTs!\n")
    return inserted_count


def main():
    """Main migration workflow"""
    print("\n" + "="*60)
    print("DREAM GPT DATABASE MIGRATION")
    print("="*60 + "\n")

    try:
        # 1. Initialize Supabase
        supabase = init_supabase()
        print("✅ Connected to Supabase\n")

        # 2. Clear existing data
        clear_existing_data(supabase)

        # 3. Read Excel and get top 50 GPTs
        top_gpts = read_excel_and_get_top_gpts(limit=50)

        # 4. Insert into Supabase
        inserted = insert_gpts_to_supabase(supabase, top_gpts)

        print("="*60)
        print("MIGRATION COMPLETE!")
        print(f"Migrated {inserted} GPTs to Quiver Supabase")
        print("="*60 + "\n")

    except Exception as e:
        print(f"\n❌ Migration failed: {e}\n")
        raise


if __name__ == "__main__":
    main()
